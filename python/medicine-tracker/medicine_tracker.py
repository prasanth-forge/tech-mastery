from openpyxl import load_workbook
from dotenv import load_dotenv
import os
import requests
from docx import Document
from datetime import date, timedelta
from googleapiclient.discovery import build
from google.oauth2 import service_account
import io

load_dotenv()


def get_entries(worksheet):
    entries = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            entries.append(
                {
                    "serial": str(row[1]),
                    "ordered": row[2],
                    "started": row[4],
                    "ended": row[5],
                }
            )
    return entries


def find_current_entry(entries):
    for entry in entries:
        if entry["started"].date() <= date.today() <= entry["ended"].date():
            return entry


def find_next_box_num(current):
    (batch, box) = current["serial"].split(".")
    match box:
        case "1" | "5":
            return f"{int(batch)}.{int(box) + 1}"
        case "6":
            return f"{int(batch) + 1}.1"


def check_alerts(entries, current):
    batch_end = current["ended"].date()
    days_remaining = (batch_end - date.today()).days
    current_entry = current["serial"].split(".")[1]

    if current_entry not in ("1", "5", "6"):
        return

    next_box_num = find_next_box_num(current)
    next_box_entry = [e for e in entries if e["serial"] == next_box_num]

    if next_box_entry and next_box_entry[0]["ordered"] is None:
        send_telegram("📋 Next batch not ordered yet")

        if days_remaining == 7:
            send_telegram("⚠️ 7 days left — time to order next batch")
            if current_entry in ("1", "5"):
                generate_foc_letter(current_entry, batch_end)
        elif days_remaining == 4:
            send_telegram("🚨 4 days left — order urgently")
            if current_entry in ("1", "5"):
                generate_foc_letter(current_entry, batch_end)
        else:
            send_telegram(f"✅ {days_remaining} days remaining until batch ends")


def send_telegram(message):
    token = os.environ.get("TELEGRAM_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    try:
        response = requests.get(
            f"https://api.telegram.org/bot{token}/sendMessage?chat_id={chat_id}&text={message}"
        )
        if response.status_code != 200:
            response.raise_for_status()
    except Exception:
        print("Unable to notify you over telegram")


def generate_letter(template_path, output_path, replacements):
    doc = Document(template_path)
    for paragraph in doc.paragraphs:
        for run in paragraph.runs:
            for key in replacements:
                if key in run.text:
                    run.text = run.text.replace(key, str(replacements[key]))
    doc.save(output_path)


def generate_foc_letter(current_entry, batch_end_date):
    if current_entry not in ("1", "5"):
        return

    replacements = {
        "{{today_date}}": date.today(),
        "{{delivery_date}}": batch_end_date - timedelta(2),
    }
    generate_letter(
        f"assets/for_{4 if current_entry == '1' else 1}_template.docx",
        f"assets/for_{4 if current_entry == '1' else 1}_{date.today()}.docx",
        replacements,
    )


def load_workbook_from_drive(file_id, creds_path):
    credentials = service_account.Credentials.from_service_account_file(
        creds_path, scopes=["https://www.googleapis.com/auth/drive.readonly"]
    )
    service = build("drive", "v3", credentials=credentials)
    return (
        service.files()
        .export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        .execute()
    )


if __name__ == "__main__":
    wb = load_workbook(
        io.BytesIO(
            load_workbook_from_drive(
                os.environ.get("MEDICINE_TRACKER_DRIVE_SHEET_ID"),
                "./assets/service_account.json",
            )
        )
    )
    ws = wb["Schedule Of Medication"]
    entries = get_entries(ws)
    current = find_current_entry(entries)
    check_alerts(entries, current)
